import asyncio
import os
import time
import aiomysql
from fastapi import FastAPI
from openpyxl.styles import Font
from starlette.responses import FileResponse
from starlette.background import BackgroundTask
import openpyxl
import datetime


class AioMysqlClient:
    def __init__(self, host, port, username, password, db_name, **kwargs):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.db_name = db_name
        self.kwargs = kwargs
        self.conn_pool = None
        self.is_connected = False
        self.lock = asyncio.Lock()

    async def init_pool(self):
        try:
            self.conn_pool = await aiomysql.create_pool(
                host=self.host,
                port=self.port,
                user=self.username,
                password=self.password,
                db=self.db_name,
                **self.kwargs
            )
            self.is_connected = True
        except:
            self.is_connected = False

        return self

    async def insert(self, sql, args=None):
        conn = await self.conn_pool.acquire()
        cur = await self.execute(conn, sql, args)
        await conn.commit()
        conn.close()
        return cur

    async def fetch_one(self, sql, args=None):
        conn = await self.conn_pool.acquire()
        cur = await self.execute(conn, sql, args)
        if cur.rowcount == 0:
            return None
        return await cur.fetchone()

    async def fetch_all(self, sql, args=None):
        conn = await self.conn_pool.acquire()
        cur = await self.execute(conn, sql, args)
        if cur.rowcount == 0:
            return []

        return await cur.fetchall()

    async def execute(self, conn, sql, args=None):
        async with conn.cursor(aiomysql.DictCursor) as cur:
            await cur.execute(sql, args)
            return cur

    # 更新数据
    async def update(self, sql, args=None):
        conn = await self.conn_pool.acquire()
        cur = await self.execute(conn, sql, args)
        await conn.commit()
        conn.close()

    # 删除数据
    async def delete(self, sql, args=None):
        conn = await self.conn_pool.acquire()
        cur = await self.execute(conn, sql, args)
        await conn.commit()
        conn.close()


db = AioMysqlClient("127.0.0.1", 3306, "root", "123456", "bill")


def get_timestamp(timestamp):
    now = str(time.time())
    new = now.replace(now[0:10], str(timestamp))
    now = float(new)
    hour = int(time.strftime('%H', time.localtime(now)))
    if hour < 4:
        yesterday = time.localtime(now - 86400)
        yesterday_4am = time.mktime(
            time.struct_time((yesterday.tm_year, yesterday.tm_mon, yesterday.tm_mday, 4, 0, 0, 0, 0, -1)))
        today_4am = time.mktime(time.struct_time(
            (time.localtime(now).tm_year, time.localtime(now).tm_mon, time.localtime(now).tm_mday, 4, 0, 0, 0, 0, -1)))
        return yesterday_4am, today_4am
    else:
        tomorrow = time.localtime(now + 86400)
        today_4am = time.mktime(time.struct_time(
            (time.localtime(now).tm_year, time.localtime(now).tm_mon, time.localtime(now).tm_mday, 4, 0, 0, 0, 0, -1)))
        tomorrow_4am = time.mktime(
            time.struct_time((tomorrow.tm_year, tomorrow.tm_mon, tomorrow.tm_mday, 4, 0, 0, 0, 0, -1)))
        return int(today_4am), int(tomorrow_4am)


async def get(group_id, timestamp, page=1):
    try:
        start, end = get_timestamp(timestamp)
        await db.init_pool()
        sql = f"select * from `data` where `group_id` = {-group_id} AND `timestamp` >= {start} AND `timestamp` < {end} AND amount != 0"
        deposit_data = await db.fetch_all(sql)
        sql2 = f"select * from `data2` where `group_id` = {-group_id} AND `timestamp` >= {start} AND `timestamp` < {end} AND rmb != 0"
        payment_data = await db.fetch_all(sql2)
        sql3 = f"select * from `groups` where group_id = {-group_id}"
        cf = await db.fetch_one(sql3)
        # kaishi = 0
        # 1 0-9
        # 2 10-19
        # 3 20-29
        # Classification
        # if cf[]

        rate = cf['rate']
        usd_rate = cf['usd_rate']
        # if usd_rate > 0:
        rate_type = list(set([i['usd_rate'] for i in deposit_data]))
        type_list = []

        for i in rate_type:
            s = 0
            for j in deposit_data:
                if j['usd_rate'] == i:
                    s = s + j['amount']

            type_list.append({i: s})

        deposit_count = len(deposit_data)
        payment_count = len(payment_data)
        deposit_total = round(sum([i['amount'] for i in deposit_data]), 2)
        payment_total = round(sum([i['usdt'] for i in payment_data]), 2)
        yingfaxia = round(sum([i['amount'] * (1 - i['rate']) for i in deposit_data]), 2)

        if usd_rate > 0:
            deposit_total_usdt = f"｜{round(sum([i['amount'] / (i['usd_rate'] if i['usd_rate'] != 0 else 1) for i in deposit_data]), 2)} USDT"
            yingfaxia_usdt = f"｜{round(sum([i['amount'] * (1 - i['rate']) / (i['usd_rate'] if i['usd_rate'] != 0 else 1) for i in deposit_data]), 2)} USDT"
            weixiafa = f"{round(float(yingfaxia_usdt.replace('｜', '').replace(' USDT', '')) - payment_total, 2)}"
        else:
            deposit_total_usdt = ""
            yingfaxia_usdt = ""
            weixiafa = round(yingfaxia - payment_total, 2)

        num = 0
        num2 = 0
        deposit_list = []
        payment_list = []
        # for i in deposit_data[(page - 1) * 10:page * 10]:
        for i in deposit_data:
            f_time = time.strftime("%H:%M:%S", time.localtime(i['timestamp']))
            f_amount = i['amount']
            f_respondent = i['respondent']
            f_operator = i['operator']
            if usd_rate > 0:
                f_calculate = f"/ {i['usd_rate']}= {round(f_amount / (i['usd_rate'] if i['usd_rate'] != 0 else 1), 2)} U"
            else:
                f_calculate = ""
            f_json = {'time': f_time, 'amount': f_amount, 'calculate': f_calculate, 'reply': f_respondent, 'operator': f_operator}
            deposit_list.append(f_json)
            num += 1
        r_json = {'type': '入款',
                  'total_amount': f"{deposit_total}{deposit_total_usdt}",
                  'transfer_amount': f"{yingfaxia}{yingfaxia_usdt}",
                  'total': deposit_count,
                  'data': deposit_list,
                  'rate': rate,
                  'usd_rate': usd_rate}
        # for z in payment_data[(page - 1) * 10:page * 10]:
        for z in payment_data:
            f_time = time.strftime("%H:%M:%S", time.localtime(z['timestamp']))
            f_amount = z['rmb']
            f_respondent = z['respondent']
            f_operator = z['operator']
            f_json = {'time': f_time, 'amount': f_amount, 'reply': f_respondent, 'operator': f_operator}
            payment_list.append(f_json)
            num2 += 1
        c_json = {'type': '下发',
                  'total_amount': f"{payment_total} USDT",
                  'transfer_amount': f"{weixiafa} USDT",
                  'total': payment_count,
                  'data': payment_list}
        if usd_rate > 0:
            return {'code': 1, 'msg': 'success', 'data': {'deposit': r_json, 'payment': c_json, 'Classification': {'data': type_list}}}
        else:
            payment_total = round(sum([i['rmb'] for i in payment_data]), 2)
            weixiafa = round(yingfaxia - payment_total, 2)
            r_json = {'type': '入款',
                      'total_amount': f"{deposit_total}",
                      'transfer_amount': f"{yingfaxia}",
                      'total': deposit_count,
                      'data': deposit_list,
                      'rate': rate,
                      'usd_rate': 0}
            c_json = {'type': '下发',
                      'total_amount': f"{payment_total}",
                      'transfer_amount': f"{weixiafa}",
                      'total': payment_count,
                      'data': payment_list}
            return {'code': 1, 'msg': 'success', 'data': {'deposit': r_json, 'payment': c_json, 'Classification': {'data': []}}}
    except Exception as e:
        return {'code': 0, 'msg': str(e)}


def generate_excel_r(data, data2, group_id):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('入账')
    ws['A1'] = '时间'
    ws['B1'] = '人民币'
    ws['C1'] = '费率'
    ws['D1'] = '汇率'
    ws['E1'] = '操作人'
    ws['F1'] = '回复人'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=datetime.datetime.fromtimestamp(int(item['timestamp'])).strftime("%H:%M:%S"))
        ws.cell(row=row, column=2, value=item['amount'])
        ws.cell(row=row, column=3, value=item['rate'])
        ws.cell(row=row, column=4, value=item['usd_rate'])
        ws.cell(row=row, column=5, value=item['operator'])
        ws.cell(row=row, column=6, value=item['respondent'])

    ws = wb.create_sheet('出账')
    ws['A1'] = '时间'
    ws['B1'] = '人民币'
    ws['C1'] = 'USD'
    ws['D1'] = '操作人'
    ws['E1'] = '回复人'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    for row, item in enumerate(data2, start=2):
        ws.cell(row=row, column=1, value=datetime.datetime.fromtimestamp(int(item['timestamp'])).strftime("%H:%M:%S"))
        ws.cell(row=row, column=2, value=item['rmb'])
        ws.cell(row=row, column=3, value=item['usdt'])
        ws.cell(row=row, column=4, value=item['operator'])
        ws.cell(row=row, column=5, value=item['respondent'])

    ws = wb['Sheet']
    wb.remove(ws)

    t = datetime.datetime.now().strftime("%Y_%m_%d")
    filename = f'{group_id}_{t}.xlsx'
    wb.save(filename)
    return filename


def generate_excel_c(data, group_id):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('出账')
    ws['A1'] = '时间'
    ws['B1'] = '人民币'
    ws['C1'] = 'USD'
    ws['D1'] = '操作人'
    ws['E1'] = '回复人'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=datetime.datetime.fromtimestamp(int(item['timestamp'])).strftime("%H:%M:%S"))
        ws.cell(row=row, column=2, value=item['rmb'])
        ws.cell(row=row, column=3, value=item['usdt'])
        ws.cell(row=row, column=4, value=item['operator'])
        ws.cell(row=row, column=5, value=item['respondent'])
    t = datetime.datetime.now().strftime("%Y_%m_%d")
    filename = f'expenses_{group_id}_{t}.xlsx'
    wb.save(filename)
    return filename


async def get_excel(group_id, timestamp):
    start, end = get_timestamp(timestamp)
    await db.init_pool()
    sql = f"select amount,timestamp,rate,usd_rate,respondent,operator from `data` where `group_id` = {-group_id} AND `timestamp` >= {start} AND `timestamp` < {end} AND amount != 0"
    deposit_data = await db.fetch_all(sql)
    sql2 = f"select rmb,usdt,timestamp,respondent,operator from `data2` where `group_id` = {-group_id} AND `timestamp` >= {start} AND `timestamp` < {end} AND rmb != 0"
    payment_data = await db.fetch_all(sql2)
    if deposit_data or payment_data:
        r = generate_excel_r(deposit_data, payment_data, group_id)
        return r


api = FastAPI()


@api.get("/bill/")
async def read_bill(group: int, timestamp: int, page: int = 1):
    if group == 0 or timestamp == 0:
        return {'code': 0, 'msg': '参数错误'}
    data = await get(group, timestamp, page)
    return data


@api.get("/download/")
async def download(group: int, timestamp: int):
    if group == 0 or timestamp == 0:
        return {'code': 0, 'msg': '参数错误'}
    excel = await get_excel(group, timestamp)
    filename = excel
    try:
        return FileResponse(
            filename,
            filename=excel,
            background=BackgroundTask(lambda: os.remove(filename)),
        )
    except Exception as e:
        return {'code': 0, 'msg': 'fail'}


if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    loop.run_until_complete(get(1001836046135, 1682438400, page=1))
